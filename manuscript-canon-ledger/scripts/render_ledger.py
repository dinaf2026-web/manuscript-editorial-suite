# -*- coding: utf-8 -*-
"""Render the living canon ledger JSON into a human-readable .xlsx view.

Usage:
    python render_ledger.py [json_path] [xlsx_out]

Defaults:
    json_path : the canon-ledger.json in the nearest .manuscript/ folder, found by
                walking up from the current working directory.
    xlsx_out  : canon-ledger.xlsx beside the JSON.

The JSON is the source of truth; this is a one-way view generator.
Requires: openpyxl  (pip install openpyxl)
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def find_default_json():
    """Walk up from CWD looking for .manuscript/canon-ledger.json."""
    d = os.path.abspath(os.getcwd())
    while True:
        candidate = os.path.join(d, ".manuscript", "canon-ledger.json")
        if os.path.isfile(candidate):
            return candidate
        parent = os.path.dirname(d)
        if parent == d:
            return candidate  # fall through to a sensible default path even if absent
        d = parent


json_path = sys.argv[1] if len(sys.argv) > 1 else find_default_json()
out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
    os.path.dirname(os.path.abspath(json_path)), "canon-ledger.xlsx")

with open(json_path, encoding="utf-8") as f:
    L = json.load(f)

HEADER = "2E3A48"; WHITE = "FFFFFF"
STATUS_FILL = {"STABLE": "D5F5E3", "DISPUTED": "FAD7D2", "EVOLVING": "FCF3CF"}
SEV_FILL = {"CRITICAL": "C0392B", "HIGH": "E67E22", "MEDIUM": "F1C40F", "LOW": "7F8C8D"}
thin = Side(style="thin", color="C9CDD2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(c):
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=HEADER)
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    c.border = border


def cel(c, size=10):
    c.font = Font(name="Calibri", size=size)
    c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    c.border = border


wb = Workbook()

# ---- Sheet: Canon State ----
ws = wb.active; ws.title = "Canon State"
cols = ["Entity", "Type", "Attribute", "Value", "Status", "Baseline / since", "Conflict / note", "Governed by"]
widths = [20, 12, 20, 46, 12, 30, 46, 12]
for i, c in enumerate(cols, 1):
    hdr(ws.cell(row=1, column=i, value=c)); ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
r = 2
for ent in L.get("canon", []):
    for attr, info in ent.get("attributes", {}).items():
        baseline = info.get("baseline") or info.get("since") or ""
        conflict = info.get("conflict") or info.get("note") or ""
        vals = [ent["entity"], ent.get("type", ""), attr, info.get("value", ""), info.get("status", ""),
                baseline, conflict, info.get("governed_by", "")]
        for i, v in enumerate(vals, 1): cel(ws.cell(row=r, column=i, value=v))
        st = ws.cell(row=r, column=5)
        if info.get("status") in STATUS_FILL:
            st.fill = PatternFill("solid", fgColor=STATUS_FILL[info["status"]])
            st.font = Font(name="Calibri", bold=True, size=10)
            st.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        r += 1
ws.freeze_panes = "A2"

# ---- Sheet: Deltas ----
ws = wb.create_sheet("Deltas (intentional)")
cols = ["ID", "Entity", "Attribute", "From", "To", "Book", "Chapter", "Reason", "Status", "Recorded"]
widths = [6, 18, 18, 34, 40, 7, 18, 46, 22, 12]
for i, c in enumerate(cols, 1):
    hdr(ws.cell(row=1, column=i, value=c)); ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
r = 2
for d in L.get("deltas", []):
    vals = [d.get("id"), d.get("entity"), d.get("attribute"), d.get("from"), d.get("to"),
            d.get("book"), d.get("chapter"), d.get("reason"), d.get("status"), d.get("recorded")]
    for i, v in enumerate(vals, 1): cel(ws.cell(row=r, column=i, value=v))
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FCF3CF")
    r += 1
ws.freeze_panes = "A2"

# ---- Sheet: Open Questions ----
ws = wb.create_sheet("Open Questions")
cols = ["ID", "Severity", "Title", "Baseline", "Conflict", "Options", "Recommended", "Decision", "Owner"]
widths = [7, 11, 30, 34, 40, 40, 40, 18, 14]
for i, c in enumerate(cols, 1):
    hdr(ws.cell(row=1, column=i, value=c)); ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
r = 2
for q in L.get("open_questions", []):
    opts = "\n".join(f"- {o}" for o in q.get("options", []))
    vals = [q.get("id"), q.get("severity"), q.get("title"), q.get("baseline"), q.get("conflict"),
            opts, q.get("recommended"), q.get("decision") or "(undecided)", q.get("owner")]
    for i, v in enumerate(vals, 1): cel(ws.cell(row=r, column=i, value=v))
    sv = ws.cell(row=r, column=2)
    if q.get("severity") in SEV_FILL:
        sv.fill = PatternFill("solid", fgColor=SEV_FILL[q["severity"]])
        sv.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        sv.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    if not q.get("decision"):
        ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor="FAD7D2")
    r += 1
ws.freeze_panes = "A2"

# ---- Sheet: About ----
ws = wb.create_sheet("About")
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 80
meta = [("Title", L.get("title", "")), ("Series", L.get("series", "")), ("Author", L.get("author", "")),
        ("Ledger version", L.get("ledger_version", "")), ("Updated", L.get("updated", "")),
        ("Baseline source", L.get("baseline_source", "")),
        ("Books", " | ".join(f"{b.get('id')}: {b.get('title')} ({b.get('status')})" for b in L.get("books", [])))]
for i, (k, v) in enumerate(meta, 1):
    a = ws.cell(row=i, column=1, value=k); b = ws.cell(row=i, column=2, value=v)
    a.font = Font(name="Calibri", bold=True, size=11, color=HEADER); cel(b, 11)
leg = ws.max_row + 2
ws.cell(row=leg, column=1, value="Status legend").font = Font(name="Calibri", bold=True, size=11, color=HEADER)
for j, (k, v) in enumerate(L.get("legend", {}).get("status", {}).items(), 1):
    ws.cell(row=leg + j, column=1, value=k).font = Font(name="Calibri", bold=True, size=10)
    cel(ws.cell(row=leg + j, column=2, value=v), 10)

wb.save(out_path)
print("Rendered:", out_path)
